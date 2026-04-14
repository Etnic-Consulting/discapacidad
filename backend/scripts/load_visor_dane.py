"""
Load all sheets from DANE Visor Excel (visor-pueblos-indigenas-06-2021.xlsx)
into PostgreSQL schema visor_dane.

Covers 23 data sheets (numbered 1-23 in the workbook).
Sheets 2-6 have data already extracted from REDATAM but are loaded anyway for completeness.
"""

import sys
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── Config ──────────────────────────────────────────────────────────────
EXCEL_PATH = r"D:\1.Programacion\1.onic-smt-dashboard\visor-pueblos-indigenas-06-2021.xlsx"
DB = dict(host="localhost", port=5450, user="smt_admin",
          password="smt_onic_2026", dbname="smt_onic")

# ── Helpers ─────────────────────────────────────────────────────────────

def clean(val):
    """Normalize a cell value: NA/None -> None, strip strings."""
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip()
        if v.upper() in ("NA", "N/A", "", "NO INFORMA"):
            return None
        return v
    return val


def clean_numeric(val):
    """Return numeric or None."""
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip()
        if v.upper() in ("NA", "N/A", "", "-", "NO INFORMA"):
            return None
        try:
            return float(v)
        except ValueError:
            return None
    return val


def clean_int(val):
    """Return int or None."""
    n = clean_numeric(val)
    if n is None:
        return None
    return int(n)


def read_sheet(wb, name, ncols, skip_none_col=None):
    """Read a sheet, return list of tuples with ncols columns.
    skip_none_col: if set, skip rows where that column index is entirely None (header-like blanks).
    """
    ws = wb[name]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        vals = list(row[:ncols])
        # Skip completely empty rows
        if all(v is None for v in vals):
            continue
        rows.append(tuple(vals))
    return rows


# ── DDL ─────────────────────────────────────────────────────────────────

DDL = """
-- Schema
CREATE SCHEMA IF NOT EXISTS visor_dane;

-- Sheet 1: Poblacion 2005 vs 2018 por pueblo
DROP TABLE IF EXISTS visor_dane.poblacion_pueblo CASCADE;
CREATE TABLE visor_dane.poblacion_pueblo (
    cod_pueblo INTEGER,
    pueblo VARCHAR(120),
    poblacion_2005 INTEGER,
    poblacion_2018 INTEGER
);

-- Sheet 2: Poblacion por clase y pueblo
DROP TABLE IF EXISTS visor_dane.poblacion_clase CASCADE;
CREATE TABLE visor_dane.poblacion_clase (
    clase VARCHAR(60),
    territorio VARCHAR(120),
    cod_pueblo INTEGER,
    poblacion_2018 INTEGER
);

-- Sheet 3: Poblacion por depto/mpio/pueblo
DROP TABLE IF EXISTS visor_dane.poblacion_depto_mpio CASCADE;
CREATE TABLE visor_dane.poblacion_depto_mpio (
    departamento VARCHAR(100),
    municipio VARCHAR(120),
    cod_pueblo INTEGER,
    poblacion_2018 INTEGER
);

-- Sheet 4: Piramide poblacional por grandes grupos de edad
DROP TABLE IF EXISTS visor_dane.poblacion_mapa CASCADE;
CREATE TABLE visor_dane.poblacion_mapa (
    cod_pueblo INTEGER,
    sexo VARCHAR(20),
    edad VARCHAR(40),
    poblacion_2018 INTEGER
);

-- Sheet 5: Piramide poblacional (sexo x edad x pueblo)
DROP TABLE IF EXISTS visor_dane.piramide_pueblo CASCADE;
CREATE TABLE visor_dane.piramide_pueblo (
    cod_pueblo INTEGER,
    sexo VARCHAR(20),
    edad VARCHAR(30),
    poblacion_2018 INTEGER
);

-- Sheet 6: Poblacion por depto/mpio/pueblo (llave)
DROP TABLE IF EXISTS visor_dane.poblacion_mpio_pueblo CASCADE;
CREATE TABLE visor_dane.poblacion_mpio_pueblo (
    cod_depto INTEGER,
    cod_mpio INTEGER,
    cod_pueblo INTEGER,
    llave VARCHAR(30),
    poblacion_2018 INTEGER
);

-- Sheet 7: Lengua nativa por pueblo
DROP TABLE IF EXISTS visor_dane.lengua_pueblo CASCADE;
CREATE TABLE visor_dane.lengua_pueblo (
    clase VARCHAR(60),
    territorio VARCHAR(120),
    cod_pueblo INTEGER,
    habla_lengua VARCHAR(30),
    entiende_lengua VARCHAR(30),
    poblacion INTEGER
);

-- Sheet 8 (3.1): Autoreconocimiento - poblacion por depto/mpio/resguardo
DROP TABLE IF EXISTS visor_dane.autoreconocimiento CASCADE;
CREATE TABLE visor_dane.autoreconocimiento (
    nivel VARCHAR(20),
    nombre VARCHAR(200),
    poblacion_2018 INTEGER,
    orden_rank NUMERIC
);

-- Sheet 9: Nivel educativo por pueblo
DROP TABLE IF EXISTS visor_dane.educacion_pueblo CASCADE;
CREATE TABLE visor_dane.educacion_pueblo (
    cod_pueblo INTEGER,
    nivel_educativo VARCHAR(120),
    sexo VARCHAR(20),
    edades VARCHAR(40),
    poblacion_2018 INTEGER
);

-- Sheet 10: Asistencia escolar
DROP TABLE IF EXISTS visor_dane.asistencia_pueblo CASCADE;
CREATE TABLE visor_dane.asistencia_pueblo (
    cod_pueblo INTEGER,
    asistencia_escolar VARCHAR(30),
    sexo VARCHAR(20),
    edades VARCHAR(40),
    poblacion_2018 INTEGER
);

-- Sheet 11: Alfabetismo por pueblo
DROP TABLE IF EXISTS visor_dane.alfabetismo_pueblo CASCADE;
CREATE TABLE visor_dane.alfabetismo_pueblo (
    cod_pueblo INTEGER,
    alfabetismo VARCHAR(30),
    sexo VARCHAR(20),
    poblacion_2018 INTEGER
);

-- Sheet 12: Funcionamiento humano (capacidades diversas) por pueblo y sexo
DROP TABLE IF EXISTS visor_dane.funcionamiento_pueblo CASCADE;
CREATE TABLE visor_dane.funcionamiento_pueblo (
    cod_pueblo INTEGER,
    sexo VARCHAR(20),
    condicion_fisica VARCHAR(30),
    poblacion_2018 INTEGER
);

-- Sheet 13: Trabajo por pueblo y sexo
DROP TABLE IF EXISTS visor_dane.trabajo_pueblo CASCADE;
CREATE TABLE visor_dane.trabajo_pueblo (
    cod_pueblo INTEGER,
    sexo VARCHAR(20),
    trabajo VARCHAR(200),
    poblacion_2018 INTEGER
);

-- Sheet 14: Servicios vivienda por pueblo (6 servicios side-by-side)
DROP TABLE IF EXISTS visor_dane.servicios_vivienda_pueblo CASCADE;
CREATE TABLE visor_dane.servicios_vivienda_pueblo (
    cod_pueblo INTEGER,
    servicio VARCHAR(60),
    tiene VARCHAR(30),
    poblacion_2018 INTEGER
);

-- Sheet 15: Tipo vivienda por pueblo
DROP TABLE IF EXISTS visor_dane.tipo_vivienda_pueblo CASCADE;
CREATE TABLE visor_dane.tipo_vivienda_pueblo (
    cod_pueblo INTEGER,
    tipo_vivienda VARCHAR(120),
    poblacion_2018 INTEGER
);

-- Sheet 16: Material pisos y paredes
DROP TABLE IF EXISTS visor_dane.material_vivienda_pueblo CASCADE;
CREATE TABLE visor_dane.material_vivienda_pueblo (
    cod_pueblo INTEGER,
    tipo_material VARCHAR(20),
    material VARCHAR(150),
    poblacion_2018 INTEGER
);

-- Sheet 17: Tamano hogar por pueblo
DROP TABLE IF EXISTS visor_dane.tamano_hogar_pueblo CASCADE;
CREATE TABLE visor_dane.tamano_hogar_pueblo (
    cod_pueblo INTEGER,
    tamano VARCHAR(40),
    poblacion_2018 INTEGER
);

-- Sheet 18: Tipo hogar por pueblo
DROP TABLE IF EXISTS visor_dane.tipo_hogar_pueblo CASCADE;
CREATE TABLE visor_dane.tipo_hogar_pueblo (
    cod_pueblo INTEGER,
    tipo_hogar VARCHAR(60),
    poblacion INTEGER
);

-- Sheet 19: Jefe hogar por sexo y pueblo
DROP TABLE IF EXISTS visor_dane.jefe_hogar_pueblo CASCADE;
CREATE TABLE visor_dane.jefe_hogar_pueblo (
    cod_pueblo INTEGER,
    sexo VARCHAR(20),
    poblacion_2018 INTEGER
);

-- Sheet 20: Resguardos x pueblos (matrix)
DROP TABLE IF EXISTS visor_dane.resguardo_pueblo CASCADE;
CREATE TABLE visor_dane.resguardo_pueblo (
    resguardo VARCHAR(200),
    cod_pueblo INTEGER,
    poblacion INTEGER
);

-- Sheet 21: NBI por pueblo (CRITICAL)
DROP TABLE IF EXISTS visor_dane.nbi_pueblo CASCADE;
CREATE TABLE visor_dane.nbi_pueblo (
    cod_pueblo INTEGER,
    pct_nbi NUMERIC(8,2),
    pct_miseria NUMERIC(8,2),
    nbi_vivienda NUMERIC(8,2),
    nbi_servicios NUMERIC(8,2),
    nbi_hacinamiento NUMERIC(8,2),
    nbi_inasistencia NUMERIC(8,2),
    nbi_dependencia NUMERIC(8,2),
    total_personas INTEGER
);

-- Sheet 22: NBI por pueblo x area
DROP TABLE IF EXISTS visor_dane.nbi_pueblo_area CASCADE;
CREATE TABLE visor_dane.nbi_pueblo_area (
    cod_pueblo INTEGER,
    area VARCHAR(10),
    pct_nbi NUMERIC(8,2),
    pct_miseria NUMERIC(8,2),
    nbi_vivienda NUMERIC(8,2),
    nbi_servicios NUMERIC(8,2),
    nbi_hacinamiento NUMERIC(8,2),
    nbi_inasistencia NUMERIC(8,2),
    nbi_dependencia NUMERIC(8,2),
    total_personas INTEGER
);

-- Sheet 23: IPM por pueblo (CRITICAL)
DROP TABLE IF EXISTS visor_dane.ipm_pueblo CASCADE;
CREATE TABLE visor_dane.ipm_pueblo (
    cod_pueblo INTEGER,
    ipm NUMERIC(8,2),
    sin_privacion NUMERIC(8,2),
    ipm_cabecera NUMERIC(8,2),
    ipm_resto NUMERIC(8,2)
);
"""


def load_all():
    print("Opening Excel workbook...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"  Sheets found: {len(wb.sheetnames)}")

    conn = psycopg2.connect(**DB)
    conn.autocommit = False
    cur = conn.cursor()

    # Create schema + tables
    print("\nCreating schema and tables...")
    cur.execute(DDL)
    conn.commit()
    print("  Done.")

    results = {}

    # ── Sheet 1: poblacion_pueblo ───────────────────────────────────────
    print("\n[Sheet 1] poblacion_pueblo...")
    rows = []
    ws = wb["1"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        pueblo = clean(row[1])
        p2005 = clean_int(row[2])
        p2018 = clean_int(row[3])
        rows.append((cod, pueblo, p2005, p2018))
    execute_values(cur, "INSERT INTO visor_dane.poblacion_pueblo VALUES %s", rows)
    conn.commit()
    results["poblacion_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 2: poblacion_clase ────────────────────────────────────────
    print("\n[Sheet 2] poblacion_clase...")
    rows = []
    ws = wb["2"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        clase = clean(row[0])
        territorio = clean(row[1])
        cod = clean_int(row[2])
        if cod is None:
            continue
        p2018 = clean_int(row[3])
        rows.append((clase, territorio, cod, p2018))
    execute_values(cur, "INSERT INTO visor_dane.poblacion_clase VALUES %s", rows)
    conn.commit()
    results["poblacion_clase"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 3: poblacion_depto_mpio ─────────────────────────────────
    # Cols: DEPARTAMENTO, MUNICIPIO, PA11_COD_ETNIA, POBLACION_2018
    print("\n[Sheet 3] poblacion_depto_mpio...")
    ws = wb["3"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        depto = clean(row[0])
        mpio = clean(row[1])
        cod = clean_int(row[2])
        if cod is None:
            continue
        pob = clean_int(row[3])
        rows.append((depto, mpio, cod, pob))
    if rows:
        batch = 5000
        for i in range(0, len(rows), batch):
            execute_values(cur, "INSERT INTO visor_dane.poblacion_depto_mpio VALUES %s", rows[i:i+batch])
        conn.commit()
    results["poblacion_depto_mpio"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 4: poblacion_mapa (broad age pyramid) ─────────────────────
    # Cols: PA11_COD_ETNIA, SEXO, EDAD, POBLACION_2018
    print("\n[Sheet 4] poblacion_mapa...")
    ws = wb["4"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        sexo = clean(row[1])
        edad = clean(row[2])
        pob = clean_int(row[3])
        rows.append((cod, sexo, edad, pob))
    if rows:
        execute_values(cur, "INSERT INTO visor_dane.poblacion_mapa VALUES %s", rows)
        conn.commit()
    results["poblacion_mapa"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 5: piramide_pueblo ────────────────────────────────────────
    print("\n[Sheet 5] piramide_pueblo...")
    rows = []
    ws = wb["5"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        sexo = clean(row[1])
        edad = clean(row[2])
        p2018 = clean_int(row[3])
        rows.append((cod, sexo, edad, p2018))
    if rows:
        batch = 5000
        for i in range(0, len(rows), batch):
            execute_values(cur, "INSERT INTO visor_dane.piramide_pueblo VALUES %s", rows[i:i+batch])
        conn.commit()
    results["piramide_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 6: poblacion_mpio_pueblo ──────────────────────────────────
    print("\n[Sheet 6] poblacion_mpio_pueblo...")
    rows = []
    ws = wb["6"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod_depto = clean_int(row[0])
        cod_mpio = clean_int(row[1])
        cod_pueblo = clean_int(row[2])
        if cod_pueblo is None:
            continue
        llave = clean(row[3])
        p2018 = clean_int(row[4])
        rows.append((cod_depto, cod_mpio, cod_pueblo, llave, p2018))
    if rows:
        batch = 5000
        for i in range(0, len(rows), batch):
            execute_values(cur, "INSERT INTO visor_dane.poblacion_mpio_pueblo VALUES %s", rows[i:i+batch])
        conn.commit()
    results["poblacion_mpio_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 7: lengua_pueblo ──────────────────────────────────────────
    # Cols: CLASE, TERRITORIO, PA11_COD_ETNIA, HABLA_LENGUA, ENTIENDE_LENGUA, POBLACION_IND
    print("\n[Sheet 7] lengua_pueblo...")
    rows = []
    ws = wb["7"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        clase = clean(row[0])
        territorio = clean(row[1])
        cod = clean_int(row[2])
        if cod is None:
            continue
        habla = clean(row[3])
        entiende = clean(row[4])
        pob = clean_int(row[5])
        rows.append((clase, territorio, cod, habla, entiende, pob))
    if rows:
        batch = 5000
        for i in range(0, len(rows), batch):
            execute_values(cur, "INSERT INTO visor_dane.lengua_pueblo VALUES %s", rows[i:i+batch])
        conn.commit()
    results["lengua_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 8 (3.1): autoreconocimiento ──────────────────────────────
    # 3 blocks side by side: DEPTO+pop, MUNICIPIO+pop, RESGUARDO+pop
    print("\n[Sheet 8/3.1] autoreconocimiento...")
    rows = []
    ws = wb["3.1"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        # Block 1: Departamento (col 0), poblacion (col 1), rank (col 2)
        depto = clean(vals[0])
        pob_d = clean_int(vals[1])
        if depto is not None and depto != "TOTAL":
            rows.append(("departamento", depto, pob_d, clean_numeric(vals[2])))
        # Block 2: Municipio (col 3), poblacion (col 4), rank (col 5)
        if len(vals) > 4:
            mpio = clean(vals[3])
            pob_m = clean_int(vals[4])
            if mpio is not None and mpio != "TOTAL":
                rows.append(("municipio", mpio, pob_m, clean_numeric(vals[5])))
        # Block 3: Resguardo (col 6), poblacion (col 7), rank (col 8)
        if len(vals) > 7:
            resg = clean(vals[6])
            pob_r = clean_int(vals[7])
            if resg is not None and resg != "TOTAL":
                rows.append(("resguardo", resg, pob_r, clean_numeric(vals[8])))
    if rows:
        execute_values(cur, "INSERT INTO visor_dane.autoreconocimiento VALUES %s", rows)
        conn.commit()
    results["autoreconocimiento"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 9: educacion_pueblo ───────────────────────────────────────
    print("\n[Sheet 9] educacion_pueblo...")
    rows = []
    ws = wb["9"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        nivel = clean(row[1])
        sexo = clean(row[2])
        edades = clean(row[3])
        p2018 = clean_int(row[4])
        rows.append((cod, nivel, sexo, edades, p2018))
    if rows:
        batch = 5000
        for i in range(0, len(rows), batch):
            execute_values(cur, "INSERT INTO visor_dane.educacion_pueblo VALUES %s", rows[i:i+batch])
        conn.commit()
    results["educacion_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 10: asistencia_pueblo ─────────────────────────────────────
    print("\n[Sheet 10] asistencia_pueblo...")
    rows = []
    ws = wb["10"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        asistencia = clean(row[1])
        sexo = clean(row[2])
        edades = clean(row[3])
        p2018 = clean_int(row[4])
        rows.append((cod, asistencia, sexo, edades, p2018))
    if rows:
        batch = 5000
        for i in range(0, len(rows), batch):
            execute_values(cur, "INSERT INTO visor_dane.asistencia_pueblo VALUES %s", rows[i:i+batch])
        conn.commit()
    results["asistencia_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 11: alfabetismo_pueblo ────────────────────────────────────
    print("\n[Sheet 11] alfabetismo_pueblo...")
    rows = []
    ws = wb["11"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        alfa = clean(row[1])
        sexo = clean(row[2])
        p2018 = clean_int(row[3])
        rows.append((cod, alfa, sexo, p2018))
    execute_values(cur, "INSERT INTO visor_dane.alfabetismo_pueblo VALUES %s", rows)
    conn.commit()
    results["alfabetismo_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 12: funcionamiento_pueblo ─────────────────────────────────
    print("\n[Sheet 12] funcionamiento_pueblo...")
    rows = []
    ws = wb["12"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        sexo = clean(row[1])
        condicion = clean(row[2])
        p2018 = clean_int(row[3])
        rows.append((cod, sexo, condicion, p2018))
    execute_values(cur, "INSERT INTO visor_dane.funcionamiento_pueblo VALUES %s", rows)
    conn.commit()
    results["funcionamiento_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 13: trabajo_pueblo ────────────────────────────────────────
    print("\n[Sheet 13] trabajo_pueblo...")
    rows = []
    ws = wb["13"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        sexo = clean(row[1])
        trabajo = clean(row[2])
        p2018 = clean_int(row[3])
        rows.append((cod, sexo, trabajo, p2018))
    if rows:
        batch = 5000
        for i in range(0, len(rows), batch):
            execute_values(cur, "INSERT INTO visor_dane.trabajo_pueblo VALUES %s", rows[i:i+batch])
        conn.commit()
    results["trabajo_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 14: servicios_vivienda_pueblo ─────────────────────────────
    # 6 services side-by-side in blocks of (cod_pueblo, servicio_val, poblacion, None)
    # Header: PA11_COD_ETNIA, ENERGIA_ELECTRICA, POB, None, PA11_COD_ETNIA, ACUEDUCTO, POB, None, ...
    print("\n[Sheet 14] servicios_vivienda_pueblo...")
    services = [
        (0, 1, 2, "Energia electrica"),
        (4, 5, 6, "Acueducto"),
        (8, 9, 10, "Alcantarillado"),
        (12, 13, 14, "Recoleccion basura"),
        (16, 17, 18, "Gas natural"),
        (20, 21, 22, "Internet"),
    ]
    rows = []
    ws = wb["14"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        for cod_idx, svc_idx, pob_idx, svc_name in services:
            if pob_idx >= len(vals):
                continue
            cod = clean_int(vals[cod_idx])
            if cod is None:
                continue
            tiene = clean(vals[svc_idx])
            pob = clean_int(vals[pob_idx])
            rows.append((cod, svc_name, tiene, pob))
    execute_values(cur, "INSERT INTO visor_dane.servicios_vivienda_pueblo VALUES %s", rows)
    conn.commit()
    results["servicios_vivienda_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 15: tipo_vivienda_pueblo ──────────────────────────────────
    print("\n[Sheet 15] tipo_vivienda_pueblo...")
    rows = []
    ws = wb["15"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        tipo = clean(row[1])
        p2018 = clean_int(row[2])
        rows.append((cod, tipo, p2018))
    execute_values(cur, "INSERT INTO visor_dane.tipo_vivienda_pueblo VALUES %s", rows)
    conn.commit()
    results["tipo_vivienda_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 16: material_vivienda_pueblo ──────────────────────────────
    # Two blocks: (cod, MATERIAL_PISOS, pob, None, cod, MATERIAL_PAREDES, pob)
    print("\n[Sheet 16] material_vivienda_pueblo...")
    rows = []
    ws = wb["16"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        # Pisos block
        cod = clean_int(vals[0])
        if cod is not None:
            mat = clean(vals[1])
            pob = clean_int(vals[2])
            rows.append((cod, "Pisos", mat, pob))
        # Paredes block
        if len(vals) > 6:
            cod2 = clean_int(vals[4])
            if cod2 is not None:
                mat2 = clean(vals[5])
                pob2 = clean_int(vals[6])
                rows.append((cod2, "Paredes", mat2, pob2))
    if rows:
        batch = 5000
        for i in range(0, len(rows), batch):
            execute_values(cur, "INSERT INTO visor_dane.material_vivienda_pueblo VALUES %s", rows[i:i+batch])
        conn.commit()
    results["material_vivienda_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 17: tamano_hogar_pueblo ───────────────────────────────────
    print("\n[Sheet 17] tamano_hogar_pueblo...")
    rows = []
    ws = wb["17"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        tamano = clean(row[1])
        p2018 = clean_int(row[2])
        rows.append((cod, tamano, p2018))
    execute_values(cur, "INSERT INTO visor_dane.tamano_hogar_pueblo VALUES %s", rows)
    conn.commit()
    results["tamano_hogar_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 18: tipo_hogar_pueblo ─────────────────────────────────────
    print("\n[Sheet 18] tipo_hogar_pueblo...")
    rows = []
    ws = wb["18"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        tipo = clean(row[1])
        pob = clean_int(row[2])
        rows.append((cod, tipo, pob))
    execute_values(cur, "INSERT INTO visor_dane.tipo_hogar_pueblo VALUES %s", rows)
    conn.commit()
    results["tipo_hogar_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 19: jefe_hogar_pueblo ─────────────────────────────────────
    print("\n[Sheet 19] jefe_hogar_pueblo...")
    rows = []
    ws = wb["19"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        sexo = clean(row[1])
        p2018 = clean_int(row[2])
        rows.append((cod, sexo, p2018))
    execute_values(cur, "INSERT INTO visor_dane.jefe_hogar_pueblo VALUES %s", rows)
    conn.commit()
    results["jefe_hogar_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 20: resguardo_pueblo (matrix) ─────────────────────────────
    print("\n[Sheet 20] resguardo_pueblo...")
    ws = wb["20"]
    # Row 1: header with pueblo codes starting col 1
    # Row 2: pueblo names
    # Row 3: column indices (2,3,4...)
    # Row 4+: resguardo name, then population per pueblo
    header_row20 = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row20 = list(row)

    pueblo_codes20 = []
    for i in range(1, len(header_row20)):
        v = clean_int(header_row20[i])
        if v is not None:
            pueblo_codes20.append((i, v))

    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):  # data starts row 4
        vals = list(row)
        resguardo = clean(vals[0])
        if resguardo is None:
            continue
        for col_idx, cod_pueblo in pueblo_codes20:
            pob = clean_int(vals[col_idx]) if col_idx < len(vals) else None
            if pob is not None and pob > 0:
                rows.append((resguardo, cod_pueblo, pob))

    if rows:
        batch = 5000
        for i in range(0, len(rows), batch):
            execute_values(cur, "INSERT INTO visor_dane.resguardo_pueblo VALUES %s", rows[i:i+batch])
        conn.commit()
    results["resguardo_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 21: nbi_pueblo (CRITICAL) ─────────────────────────────────
    print("\n[Sheet 21] nbi_pueblo (CRITICAL)...")
    rows = []
    ws = wb["21"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        pct_nbi = clean_numeric(row[1])
        pct_miseria = clean_numeric(row[2])
        nbi_viv = clean_numeric(row[3])
        nbi_ser = clean_numeric(row[4])
        nbi_hac = clean_numeric(row[5])
        nbi_ina = clean_numeric(row[6])
        nbi_dep = clean_numeric(row[7])
        total = clean_int(row[8])
        rows.append((cod, pct_nbi, pct_miseria, nbi_viv, nbi_ser, nbi_hac, nbi_ina, nbi_dep, total))
    execute_values(cur, "INSERT INTO visor_dane.nbi_pueblo VALUES %s", rows)
    conn.commit()
    results["nbi_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 22: nbi_pueblo_area ───────────────────────────────────────
    print("\n[Sheet 22] nbi_pueblo_area...")
    rows = []
    ws = wb["22"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = clean_int(row[0])
        if cod is None:
            continue
        area = clean(row[1])
        pct_nbi = clean_numeric(row[2])
        pct_miseria = clean_numeric(row[3])
        nbi_viv = clean_numeric(row[4])
        nbi_ser = clean_numeric(row[5])
        nbi_hac = clean_numeric(row[6])
        nbi_ina = clean_numeric(row[7])
        nbi_dep = clean_numeric(row[8])
        total = clean_int(row[9])
        rows.append((cod, area, pct_nbi, pct_miseria, nbi_viv, nbi_ser, nbi_hac, nbi_ina, nbi_dep, total))
    execute_values(cur, "INSERT INTO visor_dane.nbi_pueblo_area VALUES %s", rows)
    conn.commit()
    results["nbi_pueblo_area"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Sheet 23: ipm_pueblo (CRITICAL) ─────────────────────────────────
    # 3 blocks side by side: (cod, IPM, sin_privacion), (cod, IPM_cabecera), (cod, IPM_resto)
    print("\n[Sheet 23] ipm_pueblo (CRITICAL)...")
    rows = []
    ws = wb["23"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        cod = clean_int(vals[0])
        if cod is None:
            continue
        ipm = clean_numeric(vals[1])
        sin_priv = clean_numeric(vals[2])
        # Block 2: cols 4,5 = cod, ipm_cabecera
        ipm_cab = clean_numeric(vals[5]) if len(vals) > 5 else None
        # Block 3: cols 7,8 = cod, ipm_resto
        ipm_resto = clean_numeric(vals[8]) if len(vals) > 8 else None
        rows.append((cod, ipm, sin_priv, ipm_cab, ipm_resto))
    execute_values(cur, "INSERT INTO visor_dane.ipm_pueblo VALUES %s", rows)
    conn.commit()
    results["ipm_pueblo"] = len(rows)
    print(f"  Loaded {len(rows)} rows")

    # ── Summary ─────────────────────────────────────────────────────────
    wb.close()

    print("\n" + "=" * 65)
    print("  LOAD SUMMARY - visor_dane schema")
    print("=" * 65)
    total_rows = 0
    for table, count in results.items():
        print(f"  visor_dane.{table:35s} {count:>8,d} rows")
        total_rows += count
    print(f"  {'TOTAL':35s} {total_rows:>8,d} rows")
    print("=" * 65)

    # ── Verify with SQL counts ──────────────────────────────────────────
    print("\n  Verifying against PostgreSQL...")
    for table in results:
        cur.execute(f"SELECT COUNT(*) FROM visor_dane.{table}")
        db_count = cur.fetchone()[0]
        status = "OK" if db_count == results[table] else f"MISMATCH (expected {results[table]})"
        print(f"    visor_dane.{table:35s} {db_count:>8,d}  {status}")

    # Critical tables: show sample
    print("\n  NBI sample (top 5):")
    cur.execute("SELECT * FROM visor_dane.nbi_pueblo ORDER BY cod_pueblo LIMIT 5")
    for r in cur.fetchall():
        print(f"    pueblo={r[0]:>4d}  NBI={r[1]}%  miseria={r[2]}%  total={r[8]}")

    print("\n  IPM sample (top 5):")
    cur.execute("SELECT * FROM visor_dane.ipm_pueblo ORDER BY cod_pueblo LIMIT 5")
    for r in cur.fetchall():
        print(f"    pueblo={r[0]:>4d}  IPM={r[1]}%  sin_priv={r[2]}%  cab={r[3]}%  resto={r[4]}%")

    # NBI aggregate
    cur.execute("""
        SELECT COUNT(*),
               ROUND(AVG(pct_nbi), 2),
               ROUND(MIN(pct_nbi), 2),
               ROUND(MAX(pct_nbi), 2),
               SUM(total_personas)
        FROM visor_dane.nbi_pueblo
    """)
    r = cur.fetchone()
    print(f"\n  NBI aggregate: {r[0]} pueblos, avg={r[1]}%, min={r[2]}%, max={r[3]}%, total_personas={r[4]:,d}")

    cur.execute("""
        SELECT COUNT(*),
               ROUND(AVG(ipm), 2),
               ROUND(MIN(ipm), 2),
               ROUND(MAX(ipm), 2)
        FROM visor_dane.ipm_pueblo
    """)
    r = cur.fetchone()
    print(f"  IPM aggregate: {r[0]} pueblos, avg={r[1]}%, min={r[2]}%, max={r[3]}%")

    cur.close()
    conn.close()
    print("\n  Done. Connection closed.")


if __name__ == "__main__":
    load_all()
